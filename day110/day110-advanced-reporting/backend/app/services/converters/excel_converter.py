from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference
import pandas as pd
from typing import Dict, Any

class ExcelConverter:
    """Convert report data to Excel format"""
    
    def convert(self, data: pd.DataFrame, structure: Dict[str, Any], output_path: str) -> str:
        """Convert DataFrame to Excel"""
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Report Data"
        
        # Title
        ws['A1'] = structure.get("title", "Report")
        ws['A1'].font = Font(size=16, bold=True, color="2C3E50")
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:E1')
        
        # Timestamp
        ws['A2'] = f"Generated: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A2'].font = Font(size=10, italic=True)
        
        # Data starting from row 4
        start_row = 4
        
        # Headers
        for col_num, column in enumerate(data.columns, 1):
            cell = ws.cell(row=start_row, column=col_num)
            cell.value = column
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Data rows
        for row_num, row_data in enumerate(data.itertuples(index=False), start_row + 1):
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                
                if isinstance(value, (int, float)):
                    cell.value = value
                    if isinstance(value, float):
                        cell.number_format = '#,##0.00'
                else:
                    cell.value = str(value)
                
                # Alternating row colors
                if row_num % 2 == 0:
                    cell.fill = PatternFill(start_color="ECF0F1", end_color="ECF0F1", fill_type="solid")
        
        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Add chart if numeric data exists
        numeric_cols = data.select_dtypes(include=['float64', 'int64']).columns
        if len(numeric_cols) > 0:
            self._add_chart(ws, data, numeric_cols, start_row)
        
        wb.save(output_path)
        return output_path
    
    def _add_chart(self, ws, data: pd.DataFrame, numeric_cols, start_row: int):
        """Add line chart to worksheet"""
        
        chart = LineChart()
        chart.title = "Metrics Over Time"
        chart.style = 10
        chart.y_axis.title = 'Values'
        chart.x_axis.title = 'Data Points'
        
        # Add data to chart (first 50 rows, first 3 numeric columns)
        cols_to_chart = list(numeric_cols)[:3]
        
        for col in cols_to_chart:
            col_idx = data.columns.get_loc(col) + 1
            values = Reference(ws, min_col=col_idx, min_row=start_row, max_row=min(start_row + 50, start_row + len(data)))
            chart.add_data(values, titles_from_data=True)
        
        ws.add_chart(chart, "H4")
