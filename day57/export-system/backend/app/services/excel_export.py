from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from typing import List, Dict, Any
from io import BytesIO
from datetime import datetime

class ExcelExportService:
    def __init__(self):
        self.wb = Workbook(write_only=True)
        self.ws = self.wb.create_sheet("Notifications")
        self.headers_written = False
        
    def initialize(self, headers: List[str]):
        self.headers = headers
        # Write headers with formatting
        header_row = []
        for header in headers:
            header_row.append(header)
        self.ws.append(header_row)
        self.headers_written = True
        
    def write_batch(self, records: List[Dict[str, Any]]):
        for record in records:
            row = []
            for header in self.headers:
                value = record.get(header)
                if isinstance(value, datetime):
                    row.append(value.strftime('%Y-%m-%d %H:%M:%S'))
                elif value is None:
                    row.append('')
                else:
                    row.append(str(value))
            self.ws.append(row)
            
    def get_content(self) -> bytes:
        buffer = BytesIO()
        self.wb.save(buffer)
        return buffer.getvalue()
        
    def close(self):
        self.wb.close()
